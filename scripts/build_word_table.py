#!/usr/bin/env python3
"""Generate a two-column xlsx word table from a JSON file.

Usage:
    python build_word_table.py <input.json> <output.xlsx>

Input JSON format (UTF-8):
[
  {
    "word": "book",
    "entries": [
      {"pos": "n.", "meaning": "书；书本"},
      {"pos": "v.", "meaning": "预订；预约"}
    ]
  },
  ...
]

Output layout:
  Column A: word (merged vertically across its POS rows)
  Column B: one row per POS, format "<pos> <meaning>"
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADER = ["单词", "词性与释义"]
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(name="等线", size=12, bold=True, color="1F3864")
BODY_FONT = Font(name="等线", size=12)
WORD_FONT = Font(name="Calibri", size=12, bold=True)


def build(words, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "单词表"

    # Header
    for col, title in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Body
    row = 2
    for item in words:
        entries = item.get("entries") or []
        if not entries:
            entries = [{"pos": "", "meaning": item.get("meaning", "")}]
        start = row
        for e in entries:
            pos = (e.get("pos") or "").strip()
            meaning = (e.get("meaning") or "").strip()
            text = f"{pos} {meaning}".strip()
            cell = ws.cell(row=row, column=2, value=text)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row += 1
        # Word in column A, merged across its POS rows
        wcell = ws.cell(row=start, column=1, value=item.get("word", ""))
        wcell.font = WORD_FONT
        wcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if row - start > 1:
            ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        for r in range(start, row):
            ws.cell(row=r, column=1).border = BORDER

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 52

    wb.save(output_path)


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    input_path, output_path = sys.argv[1], sys.argv[2]
    data = json.loads(Path(input_path).read_text(encoding="utf-8"))
    if not isinstance(data, list):
        print("Error: input JSON must be an array of word objects", file=sys.stderr)
        sys.exit(1)
    build(data, output_path)
    total = sum(max(1, len(w.get("entries") or [])) for w in data)
    print(f"OK: {len(data)} words ({total} rows) -> {output_path}")


if __name__ == "__main__":
    main()
